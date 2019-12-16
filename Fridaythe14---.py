#!/usr/bin/env python
from openpyxl.styles import Font 
from netmiko import ConnectHandler
from testlab_device_list import device_list as devices
from datetime import datetime
import openpyxl

start_time = datetime.now()

def ip(a_device):
    data =[]
    for a_device in devices:
        net_connect = ConnectHandler(**a_device)
        output = net_connect.send_command('sh ver', use_textfsm=True)
        output = output[0]
        output0 = output.get('hostname')
        output1 = output.get('version')
        output2 = str(output.get('serial'))
        output2 = output2.replace("'", "").replace("[", "").replace("]", "")  # to remove weird formatting
        output3 = str(output.get('hardware'))
        output3 = output3.replace("'", "").replace("[", "").replace("]", "") # to remove weird formatting
        output4 = output.get('uptime')
        output5 = net_connect.send_command('sh run int vlan1 | inc ip')
        output5 = output5.replace('255.255.255.0', '').replace('ip address', '')   
        foo= [output5, output0, output1, output2, output3, output4]
        data.append(foo)
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('VLAN1')
    sheet['A1'] = 'Mgmt-intf'
    sheet['B1'] = 'Hostname'
    sheet['C1'] = 'IOS Version'
    sheet['D1'] = 'SN:'
    sheet['E1'] = 'Model'
    sheet['F1'] = 'Uptime'
    for col in ['A', 'B', 'C', 'D', 'E']:
        sheet.column_dimensions[col].width = 25
    sheet.column_dimensions['F'].width = 35
    sheet.row_dimensions[1].height = 30
    dmx = sheet.row_dimensions[1]
    dmx.font = Font(size=16, bold=True)
    start_column = 1
    for d in data[0:]:
        start_column = 1
        for r in d:
            output6 = d[0].split('.')
            start_row = (int(output6[3])) + 1
            sheet.cell(row=start_row, column=start_column).value = r
            start_column += 1 #start_row would print down/prints first item in list then +1 and so on
        wb.save('openpie.xlsx')


def main():
    ip(devices)

if __name__ == '__main__':
    main()                            

print(datetime.now() - start_time)
    
   
    
          
       
   